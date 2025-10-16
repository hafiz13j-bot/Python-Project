import os
from openpyxl import load_workbook,Workbook
from openpyxl.worksheet.table import Table,TableStyleInfo
from openpyxl.utils import get_column_letter
os.chdir(os.path.dirname(os.path.abspath(__file__)))
folder = input("Enter path to the folder containing raw data:").strip()
folder = folder.replace("\\","/")
if not os.path.exists(folder):
    print("Folder does not exist!")
merged_wb = Workbook()
merged_ws = merged_wb.active
merged_ws.title = "Merged"
header_added = False
unique_rows = set()
if not os.path.exists(folder):
    print("Folder does not exist!")
else:
    for filename in os.listdir(folder):
        if filename.endswith(".xlsx") and filename != "merged.xlsx":
            filepath = os.path.join(folder, filename)
            wb = load_workbook(filepath)
            ws = wb.active

            if not header_added:
                merged_ws.append([cell.value for cell in ws[1]])
                header_added = True
        
            for row in ws.iter_rows(min_row=2, values_only=True):
                 row_tuple = tuple(cell if cell is not None else "Unknown" for cell in row)
                 if row_tuple not in unique_rows:
                    unique_rows.add(row_tuple)
                    merged_ws.append(row_tuple)
    for col in merged_ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                length = len(str(cell.value))
                if length > max_length:
                     max_length = length
            except:
                 pass
            merged_ws.column_dimensions[col_letter].width = max_length + 2

    last_row = merged_ws.max_row
    last_col = get_column_letter(merged_ws.max_column)
    tab = Table(displayName="Mergedfile",ref=f"A1:{last_col}{last_row}")
    style = TableStyleInfo(
          name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    merged_ws.add_table(tab)

    merged_wb.save(os.path.join(folder, "merged_cleaned.xlsx"))
    print("✅Messy files merged and cleaned into merged_cleaned.xlsx") 